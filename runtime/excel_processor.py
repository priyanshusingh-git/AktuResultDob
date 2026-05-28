import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class ExcelProcessor:
    def __init__(self, logger):
        self.logger = logger

    def validate_input(self, input_path):
        self.logger.info(f"Loading input file: {input_path}")
        try:
            if not os.path.exists(input_path):
                self.logger.error("Input file not found.")
                return None

            wb = load_workbook(input_path, read_only=True, data_only=True)
            if not wb.sheetnames:
                self.logger.error("No sheets found in input file.")
                return None

            ws = wb.active

            # Find the header row (check first 10 rows for Roll No and DOB)
            header_row_idx = None
            roll_col_idx = None
            dob_col_idx = None

            for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
                row_vals_clean = [str(v).strip() if v is not None else "" for v in row]
                if "Roll No" in row_vals_clean and "DOB" in row_vals_clean:
                    header_row_idx = r_idx
                    roll_col_idx = row_vals_clean.index("Roll No")
                    dob_col_idx = row_vals_clean.index("DOB")
                    break

            if header_row_idx is None:
                self.logger.error("Input file must contain 'Roll No' and 'DOB' columns.")
                return None

            records = []
            seen_rolls = set()
            duplicate_count = 0

            # Read rows starting from header_row_idx + 1
            for r_idx, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
                if len(row) <= max(roll_col_idx, dob_col_idx):
                    continue
                roll_val = row[roll_col_idx]
                dob_val = row[dob_col_idx]

                if roll_val is None or dob_val is None:
                    continue

                from datetime import datetime, date
                def clean(v):
                    if isinstance(v, float) and v.is_integer():
                        return str(int(v)).strip()
                    if isinstance(v, (datetime, date)):
                        return v.strftime("%Y-%m-%d %H:%M:%S")
                    return str(v).strip()

                roll_str = clean(roll_val)
                dob_str = clean(dob_val)

                if not roll_str or not dob_str:
                    continue

                if roll_str in seen_rolls:
                    duplicate_count += 1
                    continue

                seen_rolls.add(roll_str)
                records.append({
                    "row_index": r_idx - header_row_idx - 1,
                    "Roll No": roll_str,
                    "DOB": dob_str
                })

            if duplicate_count > 0:
                self.logger.info(f"Removed {duplicate_count} duplicate Roll Nos from input.")

            self.logger.info(f"Successfully loaded {len(records)} valid records.")
            return records
        except Exception as e:
            self.logger.error(f"Failed to load input Excel: {str(e)}")
            return None

    def _build_row(self, student_info, attempt_data):
        row = {
            ("Info", "Name"): student_info["Name"],
            ("Info", "Roll No"): student_info["Roll No"],
        }

        for sub in attempt_data.get("Subjects", []):
            key = f"{sub['Code']} - {sub['Name']}"
            row[(key, "Internal")] = sub["Internal"]
            row[(key, "External")] = sub["External"]
            row[(key, "Total")] = sub["Total"]

        row[("Summary", "Grand Total")] = attempt_data.get("Grand Total", "")
        row[("Summary", "SGPA")] = attempt_data.get("SGPA", "")
        return row

    def export_from_checkpoint(self, checkpoint_store, run_id, output_path, semesters):
        payload = checkpoint_store.get_export_payload(run_id, semesters)
        semester_rows = payload["semesters"]
        failures = payload["failures"]

        dir_name = os.path.dirname(output_path)
        if dir_name:
            os.makedirs(dir_name, exist_ok=True)

        multiindex_sheets = []
        simple_sheets = []
        wrote_sheet = False

        try:
            wb = Workbook()

            for sem_num in semesters:
                rows = semester_rows.get(int(sem_num), [])
                if not rows:
                    continue

                built_rows = []
                for row in rows:
                    student_info = {
                        "Name": row["Name"],
                        "Roll No": row["Roll No"],
                        "DOB": row["DOB"],
                    }
                    attempt_data = {
                        "SGPA": row.get("SGPA", ""),
                        "Grand Total": row.get("Grand Total", ""),
                        "Subjects": row.get("Subjects", []),
                    }
                    built_rows.append(self._build_row(student_info, attempt_data))

                # Deduplicate rows by Roll No, keeping the last occurrence (matching pandas drop_duplicates keep='last')
                unique_built_rows = []
                seen_rolls = {}
                for idx, row_dict in enumerate(built_rows):
                    roll_no = str(row_dict.get(("Info", "Roll No"), "")).strip()
                    seen_rolls[roll_no] = idx

                last_indices = set(seen_rolls.values())
                unique_built_rows = [row_dict for idx, row_dict in enumerate(built_rows) if idx in last_indices]

                # Gather all unique subject keys in order of appearance
                subject_keys = []
                for row_dict in unique_built_rows:
                    for col_key in row_dict.keys():
                        if col_key[0] not in ("Info", "Summary"):
                            subject_key = col_key[0]
                            if subject_key not in subject_keys:
                                subject_keys.append(subject_key)

                # Order columns: Info -> Subjects -> Summary
                columns = []
                columns.append(("Info", "Name"))
                columns.append(("Info", "Roll No"))
                for subject_key in subject_keys:
                    columns.append((subject_key, "Internal"))
                    columns.append((subject_key, "External"))
                    columns.append((subject_key, "Total"))
                columns.append(("Summary", "Grand Total"))
                columns.append(("Summary", "SGPA"))

                sheet_name = f"Semester_{sem_num}"
                ws = wb.create_sheet(title=sheet_name)

                # Write headers
                # Row 1 is the group header (Info, Subject name, Summary)
                # Row 2 is the field header (Name, Roll No, Internal, External, Total, Grand Total, SGPA)
                # Column 1 (A) is the S.No index (blank in headers)
                ws.cell(row=1, column=1, value="")
                ws.cell(row=2, column=1, value="")
                for c_idx, col_key in enumerate(columns, start=2):
                    ws.cell(row=1, column=c_idx, value=col_key[0])
                    ws.cell(row=2, column=c_idx, value=col_key[1])

                # Write data rows
                for r_idx, row_dict in enumerate(unique_built_rows, start=3):
                    ws.cell(row=r_idx, column=1, value=r_idx - 2)  # S.No index
                    for c_idx, col_key in enumerate(columns, start=2):
                        ws.cell(row=r_idx, column=c_idx, value=row_dict.get(col_key, ""))

                multiindex_sheets.append(sheet_name)
                wrote_sheet = True

            if failures:
                # Deduplicate failures by Roll No, keeping the last occurrence
                seen_fails = {}
                for fail in failures:
                    roll = str(fail.get("roll_no", "")).strip()
                    seen_fails[roll] = fail
                unique_failures = list(seen_fails.values())

                sheet_name = "Failed_Students"
                ws = wb.create_sheet(title=sheet_name)

                # Write headers
                headers = ["Roll No", "Error Type", "Timestamp"]
                for c_idx, h in enumerate(headers, start=1):
                    ws.cell(row=1, column=c_idx, value=h)

                # Write data
                for r_idx, fail in enumerate(unique_failures, start=2):
                    ws.cell(row=r_idx, column=1, value=str(fail.get("roll_no", "")).strip())
                    ws.cell(row=r_idx, column=2, value=fail.get("error_type", ""))
                    ws.cell(row=r_idx, column=3, value=fail.get("timestamp", ""))

                simple_sheets.append("Failed_Students")
                wrote_sheet = True

            if not wrote_sheet:
                sheet_name = "Run_Info"
                ws = wb.create_sheet(title=sheet_name)
                ws.cell(row=1, column=1, value="Status")
                ws.cell(row=2, column=1, value="No results exported yet.")
                simple_sheets.append("Run_Info")

            # Remove default blank sheet
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

            wb.save(output_path)

            self._apply_workbook_formatting(output_path, multiindex_sheets, simple_sheets)
            self.logger.info(f"Exported report from checkpoints: {output_path}")
        except PermissionError:
            self.logger.error(f"Permission denied: {output_path} is open in another program.")
            raise
        except Exception as e:
            self.logger.error(f"Failed to export report: {str(e)}")
            raise

    def _apply_workbook_formatting(self, output_path, multiindex_sheets, simple_sheets):
        wb = load_workbook(output_path)

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        medium_side = Side(style="medium")
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        header_font = Font(bold=True)
        info_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        subject_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        summary_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        index_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        failure_fill = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")
        neutral_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        backlog_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

        for sheet_name in multiindex_sheets:
            self._format_sheet(
                wb[sheet_name],
                thin_border,
                medium_side,
                center_alignment,
                left_alignment,
                header_alignment,
                header_font,
                {
                    "info": info_fill,
                    "subject": subject_fill,
                    "summary": summary_fill,
                    "index": index_fill,
                    "failure": failure_fill,
                    "neutral": neutral_fill,
                    "backlog": backlog_fill,
                },
                is_multiindex=True,
            )

        for sheet_name in simple_sheets:
            self._format_sheet(
                wb[sheet_name],
                thin_border,
                medium_side,
                center_alignment,
                left_alignment,
                header_alignment,
                header_font,
                {
                    "info": info_fill,
                    "subject": subject_fill,
                    "summary": summary_fill,
                    "index": index_fill,
                    "failure": failure_fill,
                    "neutral": neutral_fill,
                    "backlog": backlog_fill,
                },
                is_multiindex=False,
            )

        wb.save(output_path)

    def _format_sheet(
        self,
        ws,
        thin_border,
        medium_side,
        center_alignment,
        left_alignment,
        header_alignment,
        header_font,
        fills,
        is_multiindex,
    ):
        if is_multiindex:
            self._format_multiindex_sheet(
                ws,
                thin_border,
                medium_side,
                center_alignment,
                left_alignment,
                header_alignment,
                header_font,
                fills,
            )
        else:
            self._format_simple_sheet(
                ws,
                thin_border,
                center_alignment,
                left_alignment,
                header_alignment,
                header_font,
                fills,
            )

    def _format_multiindex_sheet(
        self,
        ws,
        thin_border,
        medium_side,
        center_alignment,
        left_alignment,
        header_alignment,
        header_font,
        fills,
    ):
        self._remove_blank_header_gap(ws)
        logical_headers = self._normalize_multiindex_headers(ws)

        result_col_idx = None
        for col_idx in range(1, ws.max_column + 1):
            value = str(ws.cell(row=2, column=col_idx).value or "")
            if value == "Result":
                result_col_idx = col_idx
                break

        ws.freeze_panes = "D3"

        for col_idx in range(1, ws.max_column + 1):
            group_name = str(logical_headers[col_idx] or "")
            lower_group = group_name.lower()
            if col_idx == 1:
                fill = fills["index"]
            elif lower_group == "info":
                fill = fills["info"]
            elif lower_group == "summary":
                fill = fills["summary"]
            else:
                fill = fills["subject"]

            for row_idx in (1, 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.fill = fill
                cell.font = header_font
                cell.alignment = header_alignment if row_idx == 1 else center_alignment

        for row_idx in range(3, ws.max_row + 1):
            is_backlog = False
            if result_col_idx:
                result_value = str(ws.cell(row=row_idx, column=result_col_idx).value or "")
                if "PCP" in result_value or ("CP(" in result_value and "CP(0)" not in result_value):
                    is_backlog = True

            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                header_name = str(ws.cell(row=2, column=col_idx).value or "").strip().lower()
                if col_idx == 1:
                    cell.fill = fills["index"]
                    cell.alignment = center_alignment
                elif header_name in {"name", "roll no"}:
                    cell.alignment = left_alignment
                else:
                    cell.alignment = center_alignment

                if is_backlog:
                    cell.fill = fills["backlog"]

        self._apply_group_boundaries(ws, logical_headers, medium_side)
        self._set_multiindex_widths(ws, logical_headers)
        if ws.max_row >= 2:
            ws.auto_filter.ref = f"A2:{get_column_letter(ws.max_column)}{ws.max_row}"

    def _format_simple_sheet(
        self,
        ws,
        thin_border,
        center_alignment,
        left_alignment,
        header_alignment,
        header_font,
        fills,
    ):
        ws.freeze_panes = "A2"
        header_fill = fills["failure"] if ws.title == "Failed_Students" else fills["neutral"]

        for cell in ws[1]:
            cell.border = thin_border
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                header_name = str(ws.cell(row=1, column=cell.column).value or "").strip().lower()
                if any(token in header_name for token in ("name", "roll", "error", "timestamp", "status")):
                    cell.alignment = left_alignment
                else:
                    cell.alignment = center_alignment

        self._set_simple_widths(ws)
        if ws.max_row >= 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    def _remove_blank_header_gap(self, ws):
        if ws.max_row < 3:
            return
        if all(ws.cell(row=3, column=col_idx).value in (None, "") for col_idx in range(1, ws.max_column + 1)):
            ws.delete_rows(3, 1)

    def _normalize_multiindex_headers(self, ws):
        logical_headers = [None] * (ws.max_column + 1)
        for col_idx in range(1, ws.max_column + 1):
            logical_headers[col_idx] = ws.cell(row=1, column=col_idx).value

        row1_merges = [merged for merged in list(ws.merged_cells.ranges) if merged.min_row == 1 and merged.max_row == 1]
        for merged in row1_merges:
            top_left_value = ws.cell(row=1, column=merged.min_col).value
            for col_idx in range(merged.min_col, merged.max_col + 1):
                logical_headers[col_idx] = top_left_value
            ws.unmerge_cells(str(merged))

        for col_idx in range(2, ws.max_column + 1):
            ws.cell(row=1, column=col_idx).value = logical_headers[col_idx]

        start_col = 2
        current_value = logical_headers[2] if ws.max_column >= 2 else None
        for col_idx in range(2, ws.max_column + 2):
            next_value = logical_headers[col_idx] if col_idx <= ws.max_column else None
            if next_value != current_value:
                if current_value and col_idx - start_col > 1:
                    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=col_idx - 1)
                start_col = col_idx
                current_value = next_value

        return logical_headers

    def _apply_group_boundaries(self, ws, logical_headers, medium_side):
        for col_idx in range(2, ws.max_column + 1):
            if logical_headers[col_idx] != logical_headers[col_idx - 1]:
                for row_idx in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = Border(
                        left=medium_side,
                        right=cell.border.right,
                        top=cell.border.top,
                        bottom=cell.border.bottom,
                    )

    def _set_multiindex_widths(self, ws, logical_headers):
        max_widths = {}
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            for row_idx in range(2, min(ws.max_row, 500) + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is not None:
                    max_len = max(max_len, len(str(value)))
            max_widths[col_idx] = min(max_len + 3, 50)

        ws.column_dimensions["A"].width = 6
        for col_idx in range(2, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = max_widths[col_idx]

        current_value = logical_headers[2] if ws.max_column >= 2 else None
        start_col = 2
        for col_idx in range(2, ws.max_column + 2):
            next_value = logical_headers[col_idx] if col_idx <= ws.max_column else None
            if next_value != current_value:
                if current_value:
                    group_width = sum(max_widths[c] for c in range(start_col, col_idx))
                    needed_width = len(str(current_value)) + 6
                    if group_width < needed_width:
                        extra = (needed_width - group_width) / max(1, col_idx - start_col)
                        for inner_col in range(start_col, col_idx):
                            column_letter = get_column_letter(inner_col)
                            ws.column_dimensions[column_letter].width = min(max_widths[inner_col] + extra, 50)
                start_col = col_idx
                current_value = next_value

    def _set_simple_widths(self, ws):
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            for row_idx in range(1, min(ws.max_row, 500) + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is not None:
                    max_len = max(max_len, len(str(value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 50)
